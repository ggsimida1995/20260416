from __future__ import annotations

from openpyxl import load_workbook

from src.config import SUCCESS_FIELD_MAPPING, SUCCESS_SHEET_NAME
from src.models import AppendResult


def build_success_row(source_fields: dict[str, object]) -> dict[str, object]:
    row: dict[str, object] = {}
    for source_name, target_name in SUCCESS_FIELD_MAPPING.items():
        value = source_fields.get(source_name)
        if value in (None, ""):
            continue
        row[target_name] = value
    return row


def append_success_row(workbook_path, row_data: dict[str, object]) -> AppendResult:
    workbook = load_workbook(workbook_path)
    sheet = workbook[SUCCESS_SHEET_NAME]
    headers = {
        str(sheet.cell(1, column).value): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(1, column).value is not None
    }

    project_code_column = headers.get("项目编码")
    project_code = row_data.get("项目编码")
    if project_code_column and project_code not in (None, ""):
        for row_index in range(2, sheet.max_row + 1):
            if sheet.cell(row_index, project_code_column).value == project_code:
                return AppendResult(status="duplicate")

    target_row = sheet.max_row + 1
    for header, value in row_data.items():
        column = headers.get(header)
        if column is None:
            continue
        sheet.cell(target_row, column).value = value

    workbook.save(workbook_path)
    return AppendResult(status="appended", appended_row_index=target_row)
