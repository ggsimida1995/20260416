from __future__ import annotations

from importlib import import_module

from openpyxl import load_workbook

from src.models import ProjectExtraction


def _clean_field_name(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", "").strip()


def read_close_sheet(path) -> ProjectExtraction:
    if str(path).lower().endswith(".xls"):
        return _read_xls_close_sheet(path)

    workbook = load_workbook(path, data_only=True)

    for sheet in workbook.worksheets:
        if _sheet_has_field_header(sheet):
            raw_fields = _read_field_rows(sheet)
            return ProjectExtraction(
                raw_fields=raw_fields,
                normalized_fields=dict(raw_fields),
            )

    return ProjectExtraction()


def _read_xls_close_sheet(path) -> ProjectExtraction:
    workbook = _open_xls_workbook(path)
    for sheet in workbook.sheets():
        if _xls_sheet_has_field_header(sheet):
            raw_fields = _read_xls_field_rows(sheet)
            return ProjectExtraction(
                raw_fields=raw_fields,
                normalized_fields=dict(raw_fields),
            )
    return ProjectExtraction()


def _open_xls_workbook(path):
    try:
        xlrd = import_module("xlrd")
    except ModuleNotFoundError as exc:
        raise RuntimeError("读取 .xls 文件需要安装 xlrd，请重新安装 requirements.txt 依赖。") from exc
    return xlrd.open_workbook(path)


def _sheet_has_field_header(sheet) -> bool:
    for row in range(1, min(sheet.max_row, 10) + 1):
        left = _clean_field_name(sheet.cell(row, 1).value)
        right = _clean_field_name(sheet.cell(row, 2).value)
        if left == "字段名称" and right == "内容":
            return True
    return False


def _xls_sheet_has_field_header(sheet) -> bool:
    for row in range(min(sheet.nrows, 10)):
        left = _clean_field_name(sheet.cell_value(row, 0) if sheet.ncols > 0 else None)
        right = _clean_field_name(sheet.cell_value(row, 1) if sheet.ncols > 1 else None)
        if left == "字段名称" and right == "内容":
            return True
    return False


def _read_field_rows(sheet) -> dict[str, object]:
    raw_fields: dict[str, object] = {}
    for row in range(1, sheet.max_row + 1):
        field_name = _clean_field_name(sheet.cell(row, 1).value)
        if not field_name or field_name in {"项目关闭登记表", "字段名称"}:
            continue
        if field_name.startswith("注："):
            continue
        raw_fields[field_name] = sheet.cell(row, 2).value
    return raw_fields


def _read_xls_field_rows(sheet) -> dict[str, object]:
    raw_fields: dict[str, object] = {}
    for row in range(sheet.nrows):
        field_name = _clean_field_name(sheet.cell_value(row, 0) if sheet.ncols > 0 else None)
        if not field_name or field_name in {"项目关闭登记表", "字段名称"}:
            continue
        if field_name.startswith("注："):
            continue
        raw_fields[field_name] = _xls_cell_value(sheet, row, 1)
    return raw_fields


def _xls_cell_value(sheet, row: int, column: int) -> object:
    if column >= sheet.ncols:
        return None
    value = sheet.cell_value(row, column)
    if value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return value
