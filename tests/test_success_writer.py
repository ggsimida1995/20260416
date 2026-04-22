from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.writers.success_writer import append_success_row, build_success_row


def test_build_success_row_only_includes_non_empty_mapped_fields():
    row = build_success_row(
        {
            "项目编号": "BHE-25030367/01",
            "项目全称": "示例项目",
            "核实人": None,
        }
    )

    assert row["项目编码"] == "BHE-25030367/01"
    assert row["项目全称"] == "示例项目"
    assert "核实人" not in row


def build_success_workbook_fixture(tmp_path: Path, existing_codes: list[str]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "登记表"
    sheet.cell(1, 4).value = "项目编码"
    sheet.cell(1, 5).value = "项目全称"
    for index, code in enumerate(existing_codes, start=2):
        sheet.cell(index, 4).value = code
        sheet.cell(index, 5).value = f"项目{index}"
    path = tmp_path / "success.xlsx"
    workbook.save(path)
    return path


def test_append_success_row_skips_existing_project_code(tmp_path: Path):
    workbook_path = build_success_workbook_fixture(tmp_path, existing_codes=["BHE-25030367/01"])

    result = append_success_row(workbook_path, {"项目编码": "BHE-25030367/01"})

    assert result.status == "duplicate"


def test_append_success_row_writes_to_last_row(tmp_path: Path):
    workbook_path = build_success_workbook_fixture(tmp_path, existing_codes=["AAA"])

    result = append_success_row(workbook_path, {"项目编码": "BBB", "项目全称": "示例项目"})

    assert result.status == "appended"
    assert result.appended_row_index == 3

    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["登记表"]
    assert sheet.cell(3, 4).value == "BBB"
    assert sheet.cell(3, 5).value == "示例项目"
