import threading
import time
from pathlib import Path

import fitz
from docx import Document
from openpyxl import Workbook

from src import workflow
from src.config_store import AISettings
from src.models import AppendResult
from src.process_state import load_processed_projects, mark_project_processed
from src.workflow import run_workflow


def build_project_tree(tmp_path: Path) -> Path:
    file_root = tmp_path / "file"
    project_dir = file_root / "BHE-25030367-01"
    success_dir = file_root / "success"
    error_dir = file_root / "error"
    other_dir = file_root / "other"
    project_dir.mkdir(parents=True)
    success_dir.mkdir(parents=True)
    error_dir.mkdir(parents=True)
    other_dir.mkdir(parents=True)

    build_close_sheet(project_dir / "BHE-25030367-01项目关闭移交登记表.xlsx")
    build_docx(project_dir / "BHE-25030367-01项目竣工总结报告.docx")
    build_pdf(project_dir / "BHE-25030367-01PA竣工验收报告.pdf", signer_name="王某某")
    build_success_workbook(success_dir / "2026年关闭满意度回访表0331.xlsx")

    return file_root


def build_close_sheet(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet2"
    sheet["A2"] = "字段名称"
    sheet["B2"] = "内容"
    sheet["A5"] = "项目编号"
    sheet["B5"] = "BHE-25030367/01"
    sheet["A6"] = "项目全称"
    sheet["B6"] = "示例项目"
    sheet["A10"] = "合同额（万元）"
    sheet["B10"] = 16.3
    sheet["A20"] = "用户联系人"
    sheet["B20"] = "黄汉民"
    sheet["A22"] = "用户联系方式"
    sheet["B22"] = "14714691425"
    workbook.save(path)


def build_docx(path: Path) -> None:
    document = Document()
    document.add_paragraph(
        "项目编号：BHE-25030367/01 项目全称 示例项目 用户姓名 黄汉民 "
        "联系电话14714691425 竣工验收2026年04月09日2026年04月10日"
    )
    document.save(path)


def build_pdf(path: Path, signer_name: str) -> None:
    document = fitz.open()
    page = document.new_page()
    page.insert_text((72, 72), f"签字人姓名 {signer_name} 联系电话 14714691425 签字时间 2026-04-10")
    document.save(path)
    document.close()


def build_success_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "登记表"
    headers = ["序号", "问题是否处理", "关闭移交资料问题项", "项目编码", "项目全称"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(1, index).value = header
    workbook.save(path)


def test_run_workflow_processes_one_project_and_writes_error_on_failure(tmp_path: Path):
    file_root = build_project_tree(tmp_path)

    result = run_workflow(file_root=file_root, username="user1", password="pass1")

    assert result.failed_count == 1
    assert result.appended_count == 0
    assert result.success_project_codes == []
    assert result.error_project_codes == ["BHE-25030367/01"]
    assert result.success_workbook_path == file_root / "success" / "2026年关闭满意度回访表0331.xlsx"
    assert result.error_report_paths == [file_root / "error" / "BHE-25030367-01.txt"]
    assert (file_root / "error" / "BHE-25030367-01.txt").exists()


def test_run_workflow_emits_grouped_logs_and_persists_log_file(tmp_path: Path):
    file_root = build_project_tree(tmp_path)
    messages: list[str] = []

    result = run_workflow(
        file_root=file_root,
        username="user1",
        password="pass1",
        log_callback=messages.append,
    )

    joined = "\n".join(messages)

    assert "[PDF识别]" in joined
    assert "[BHE-25030367/01------------开始------BHE-25030367/01]" in joined
    assert "[项目] BHE-25030367-01" in joined
    assert "[xlsx字段] 项目编号 = BHE-25030367/01" in joined
    assert "[docx字段] 项目编号 = BHE-25030367/01" in joined
    assert "[pdf字段] 签字人姓名 =" in joined
    assert "[字段比对] 用户联系人" in joined
    assert "[比对结果] 失败" in joined
    assert "[BHE-25030367/01------------结束------BHE-25030367/01]" in joined
    assert result.log_path is not None
    assert result.log_path.exists()
    assert result.log_path.parent == file_root / "error" / "logs"
    log_text = result.log_path.read_text(encoding="utf-8")
    assert "[BHE-25030367/01------------开始------BHE-25030367/01]" in log_text
    assert "[xlsx字段] 项目编号 = BHE-25030367/01" in log_text
    assert "[pdf字段] 签字人姓名 =" in log_text
    assert "[BHE-25030367/01------------结束------BHE-25030367/01]" in log_text


def test_run_workflow_logs_remote_recognition_provider(monkeypatch, tmp_path: Path):
    file_root = build_project_tree(tmp_path)
    messages: list[str] = []
    monkeypatch.setattr(workflow, "read_pdf", lambda path, ai_settings=None: workflow.PdfData())

    run_workflow(
        file_root=file_root,
        username="user1",
        password="pass1",
        log_callback=messages.append,
        ai_settings=AISettings(
            enabled=True,
            ai_base_url="https://example.com/api/v3",
            ai_api_key="secret-ai-key",
            ai_model="vision-model",
            ocr_base_url="https://example.com/ocr",
            ocr_api_key="secret-ocr-key",
            request_timeout_seconds=18,
            image_max_kb=96,
        ),
    )

    joined = "\n".join(messages)

    assert "优先=AI" in joined
    assert "AI=已配置" in joined
    assert "OCR=已配置" in joined
    assert "图片上限=96KB" in joined


def test_apply_prepared_result_records_duplicate_and_writes_error_report(monkeypatch, tmp_path: Path):
    messages: list[str] = []

    class FakeLogger:
        def log(self, message: str) -> None:
            messages.append(message)

    prepared = workflow.PreparedProjectResult(
        project_name="BHE-25030367-01",
        project_code="BHE-25030367/01",
        status="compare_passed",
        log_lines=["[项目] BHE-25030367-01"],
        success_row={"项目编码": "BHE-25030367/01", "项目全称": "示例项目"},
    )
    result = workflow.WorkflowResult()
    success_workbook_path = tmp_path / "success" / "2026年关闭满意度回访表0331.xlsx"
    error_root = tmp_path / "error"

    monkeypatch.setattr(
        workflow,
        "append_success_row",
        lambda workbook_path, row_data: AppendResult(status="duplicate"),
    )

    workflow._apply_prepared_result(
        logger=FakeLogger(),
        prepared=prepared,
        result=result,
        success_workbook_path=success_workbook_path,
        error_root=error_root,
    )

    assert result.appended_count == 0
    assert result.duplicate_count == 1
    assert result.failed_count == 0
    assert result.success_project_names == []
    assert result.success_project_codes == []
    assert result.error_project_codes == ["BHE-25030367/01"]
    assert result.error_report_paths == [error_root / "BHE-25030367-01.txt"]
    assert "[BHE-25030367/01------------开始------BHE-25030367/01]" in messages
    assert "[写入成功台账] 检测到重复项目编码，跳过追加" in messages
    assert "[BHE-25030367/01------------结束------BHE-25030367/01]" in messages
    assert (error_root / "BHE-25030367-01.txt").exists() is True


def test_run_workflow_prepares_multiple_projects_in_parallel(monkeypatch, tmp_path: Path):
    file_root = tmp_path / "file"
    (file_root / "success").mkdir(parents=True)
    (file_root / "error").mkdir(parents=True)
    projects = [file_root / "P1", file_root / "P2", file_root / "P3"]
    for project_dir in projects:
        project_dir.mkdir(parents=True)

    active = 0
    max_active = 0
    lock = threading.Lock()

    def fake_prepare(project_dir: Path):
        nonlocal active, max_active
        with lock:
            active += 1
            max_active = max(max_active, active)
        time.sleep(0.05)
        with lock:
            active -= 1
        return workflow.PreparedProjectResult(
            project_name=project_dir.name,
            project_code=project_dir.name,
            status="compare_passed",
            log_lines=[f"[项目] {project_dir.name}"],
            success_row={"项目编码": project_dir.name, "项目全称": project_dir.name},
        )

    monkeypatch.setattr(workflow, "discover_projects", lambda root: projects)
    monkeypatch.setattr(workflow, "_prepare_project_result", fake_prepare)
    monkeypatch.setattr(
        workflow,
        "append_success_row",
        lambda workbook_path, row_data: AppendResult(status="appended", appended_row_index=2),
    )

    result = run_workflow(file_root=file_root, username="user1", password="pass1")

    assert max_active >= 2
    assert result.appended_count == 3


def test_run_batch_workflow_skips_projects_already_in_processed_state(tmp_path: Path):
    file_root = tmp_path / "file"
    success_dir = file_root / "success"
    error_dir = file_root / "error"
    success_dir.mkdir(parents=True)
    error_dir.mkdir(parents=True)
    project_a = file_root / "BHE-25030367-01"
    project_b = file_root / "BHE-25030368-01"
    project_a.mkdir(parents=True)
    project_b.mkdir(parents=True)

    processed_path = tmp_path / "config" / "processed_projects.json"
    mark_project_processed(processed_path, project_code=project_a.name, processed_at="2026-04-17T16:20:00+08:00")

    seen_web_projects: list[str] = []
    seen_compare_projects: list[str] = []

    def fake_web_phase_runner(*, project_dirs, **kwargs):
        seen_web_projects.extend(project_dir.name for project_dir in project_dirs)
        return workflow.WebPhaseResult(processed_projects=[project_dir.name for project_dir in project_dirs])

    def fake_compare_runner(*, project_dirs, **kwargs):
        seen_compare_projects.extend(project_dir.name for project_dir in project_dirs)
        return workflow.WorkflowResult()

    result = workflow.run_batch_workflow(
        file_root=file_root,
        username="user1",
        password="pass1",
        processed_projects_path=processed_path,
        web_phase_runner=fake_web_phase_runner,
        compare_runner=fake_compare_runner,
    )

    assert seen_web_projects == ["BHE-25030368-01"]
    assert seen_compare_projects == ["BHE-25030368-01"]
    assert result.skipped_processed_count == 1


def test_run_batch_workflow_starts_compare_only_after_web_phase_finishes(tmp_path: Path):
    file_root = tmp_path / "file"
    (file_root / "success").mkdir(parents=True)
    (file_root / "error").mkdir(parents=True)
    project_dir = file_root / "BHE-25030367-01"
    project_dir.mkdir(parents=True)

    call_order: list[str] = []

    def fake_web_phase_runner(*, project_dirs, **kwargs):
        assert [project.name for project in project_dirs] == ["BHE-25030367-01"]
        call_order.append("web:start")
        call_order.append("web:finish")
        return workflow.WebPhaseResult(processed_projects=["BHE-25030367-01"])

    def fake_compare_runner(*, project_dirs, **kwargs):
        assert [project.name for project in project_dirs] == ["BHE-25030367-01"]
        call_order.append("compare:start")
        call_order.append("compare:finish")
        return workflow.WorkflowResult()

    workflow.run_batch_workflow(
        file_root=file_root,
        username="user1",
        password="pass1",
        processed_projects_path=tmp_path / "config" / "processed_projects.json",
        web_phase_runner=fake_web_phase_runner,
        compare_runner=fake_compare_runner,
    )

    assert call_order == ["web:start", "web:finish", "compare:start", "compare:finish"]


def test_run_batch_workflow_deletes_successful_projects_and_marks_processed(tmp_path: Path):
    file_root = tmp_path / "file"
    (file_root / "success").mkdir(parents=True)
    (file_root / "error").mkdir(parents=True)
    project_dir = file_root / "BHE-25030367-01"
    project_dir.mkdir(parents=True)
    processed_path = tmp_path / "config" / "processed_projects.json"

    def fake_web_phase_runner(*, project_dirs, **kwargs):
        return workflow.WebPhaseResult(processed_projects=[project.name for project in project_dirs])

    def fake_compare_runner(*, project_dirs, **kwargs):
        return workflow.WorkflowResult(appended_count=1, success_project_names=[project_dirs[0].name])

    result = workflow.run_batch_workflow(
        file_root=file_root,
        username="user1",
        password="pass1",
        processed_projects_path=processed_path,
        web_phase_runner=fake_web_phase_runner,
        compare_runner=fake_compare_runner,
    )

    assert result.cleaned_count == 1
    assert project_dir.exists() is False
    processed = load_processed_projects(processed_path)
    assert project_dir.name in processed


def test_run_batch_workflow_keeps_project_unprocessed_when_delete_fails(monkeypatch, tmp_path: Path):
    file_root = tmp_path / "file"
    (file_root / "success").mkdir(parents=True)
    (file_root / "error").mkdir(parents=True)
    project_dir = file_root / "BHE-25030367-01"
    project_dir.mkdir(parents=True)
    processed_path = tmp_path / "config" / "processed_projects.json"

    def fake_web_phase_runner(*, project_dirs, **kwargs):
        return workflow.WebPhaseResult(processed_projects=[project.name for project in project_dirs])

    def fake_compare_runner(*, project_dirs, **kwargs):
        return workflow.WorkflowResult(appended_count=1, success_project_names=[project_dirs[0].name])

    monkeypatch.setattr(workflow, "_delete_project_dir", lambda project_dir: (_ for _ in ()).throw(OSError("delete failed")))

    result = workflow.run_batch_workflow(
        file_root=file_root,
        username="user1",
        password="pass1",
        processed_projects_path=processed_path,
        web_phase_runner=fake_web_phase_runner,
        compare_runner=fake_compare_runner,
    )

    assert result.cleaned_count == 0
    assert project_dir.exists() is True
    assert load_processed_projects(processed_path) == {}


def test_run_batch_workflow_emits_stage_logs(tmp_path: Path):
    file_root = tmp_path / "file"
    (file_root / "success").mkdir(parents=True)
    (file_root / "error").mkdir(parents=True)
    project_dir = file_root / "BHE-25030367-01"
    project_dir.mkdir(parents=True)
    messages: list[str] = []

    def fake_web_phase_runner(*, project_dirs, processed_project_codes=None, **kwargs):
        assert processed_project_codes == set()
        return workflow.WebPhaseResult(processed_projects=[project.name for project in project_dirs])

    def fake_compare_runner(*, project_dirs, **kwargs):
        return workflow.WorkflowResult(success_project_names=[])

    workflow.run_batch_workflow(
        file_root=file_root,
        username="user1",
        password="pass1",
        processed_projects_path=tmp_path / "config" / "processed_projects.json",
        web_phase_runner=fake_web_phase_runner,
        compare_runner=fake_compare_runner,
        log_callback=messages.append,
    )

    joined = "\n".join(messages)
    assert "[网页阶段] 开始" in joined
    assert "[网页阶段] 项目数: 1" in joined
    assert "[本地比对阶段] 开始" in joined
    assert "[清理阶段] 开始" in joined


def test_run_batch_workflow_default_web_phase_rediscovers_downloaded_projects(monkeypatch, tmp_path: Path):
    file_root = tmp_path / "file"
    (file_root / "success").mkdir(parents=True)
    (file_root / "error").mkdir(parents=True)
    downloaded_project = file_root / "BHE-25030367-01"

    def fake_compare_runner(*, project_dirs, **kwargs):
        assert [project.name for project in project_dirs] == ["BHE-25030367-01"]
        return workflow.WorkflowResult(success_project_names=[])

    def fake_download_batch(*, output_root, skip_project_codes=None, log_callback=None):
        downloaded_project.mkdir(parents=True)
        return {
            "saved_project_dirs": [str(downloaded_project)],
            "skipped_projects": [],
            "categories": [],
            "errors": [],
        }

    monkeypatch.setattr(workflow, "run_hollysys_download_batch", fake_download_batch)

    result = workflow.run_batch_workflow(
        file_root=file_root,
        username="user1",
        password="pass1",
        processed_projects_path=tmp_path / "config" / "processed_projects.json",
        compare_runner=fake_compare_runner,
    )

    assert result.web_processed_count == 1


def test_run_download_workflow_uses_direct_download_summary(monkeypatch, tmp_path: Path):
    file_root = tmp_path / "file"
    messages: list[str] = []
    processed_codes = {"BHE-25040404-01"}
    captured: dict[str, object] = {}

    def fake_download_batch(*, output_root, skip_project_codes=None, log_callback=None):
        captured["output_root"] = output_root
        captured["skip_project_codes"] = skip_project_codes
        if log_callback is not None:
            log_callback("[网页阶段] 分类待办: 项目关闭工作流 | 2")
        return {
            "saved_project_dirs": [
                str(file_root / "BHE-25080117-01"),
                str(file_root / "LHE-25090012-B1"),
            ],
            "skipped_projects": [
                {"project_dir_name": "BHE-25040404-01"},
            ],
            "categories": [],
            "errors": [],
        }

    monkeypatch.setattr(workflow, "run_hollysys_download_batch", fake_download_batch)

    result = workflow.run_download_workflow(
        file_root=file_root,
        username="user1",
        password="pass1",
        log_callback=messages.append,
        processed_projects_path=tmp_path / "config" / "processed_projects.json",
    )

    assert result.processed_projects == ["BHE-25080117-01", "LHE-25090012-B1"]
    assert result.skipped_projects == ["BHE-25040404-01"]
    assert captured["output_root"] == file_root
    assert captured["skip_project_codes"] == set()
    joined = "\n".join(messages)
    assert "[网页阶段] 分类待办: 项目关闭工作流 | 2" in joined
